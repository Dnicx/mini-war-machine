import csv
import sys
from openpyxl import load_workbook

'''
This script is created to read wahapedia xlsx file but it works for any xlsx file
'''

def xlsx_to_csv(xlsx_file, csv_file):
    """Convert an Excel file to CSV format."""
    try:
        workbook = load_workbook(xlsx_file)

        for sheetName in workbook.sheetnames:
            sheet = workbook[sheetName]
            
            with open(f'{csv_file}_{sheetName}.csv', 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                for row in sheet.iter_rows():
                    row_data = []
                    for cell in row:
                        if cell.hyperlink:
                            row_data.append(cell.hyperlink.target)
                        else:
                            row_data.append(cell.value)
                    writer.writerow(row_data)
            
            print(f"Successfully converted {xlsx_file} to {csv_file}")

    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python xlsx_to_csv.py <input.xlsx> <output.csv>")
        sys.exit(1)
    
    xlsx_file = sys.argv[1]
    csv_file = sys.argv[2]
    xlsx_to_csv(xlsx_file, csv_file)
